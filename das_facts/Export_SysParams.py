# Queries the SYS_PARAMS table of a specified database
# Exports the results to a spreadsheet to generate the DasFactsSheet

import sys, os, openpyxl, pymssql, datetime
from openpyxl.utils import get_column_letter

database_name = input('Enter the name of the Sys1 database (Without Sys1): ')

# Database Connection Configuration
connection = pymssql.connect(
                server=f"{os.getlogin()}\\SQLEXPRESS",
                port=1433,
                user="sa",
                password="sa123",
                database=(f'{database_name}Sys1'))

# File Location Configuration. 
prefix_path_parent = f'C:\\Users\\{os.getlogin()}\\Desktop\\Generated_Scripts\\' # Location of all generated files. 
prefix_path = f'C:\\Users\\{os.getlogin()}\\Desktop\\Generated_Scripts\\{database_name}\\' # Location of das facts sheet file. 

try: # Attempt to create directory. Continues script if directory already exists.
    os.mkdir(prefix_path_parent)
except FileExistsError:
    print(f'Directory {prefix_path_parent} already exists, continuing script.')

try: # Attempt to create directory. Continues script if directory already exists.
    os.mkdir(prefix_path)
except FileExistsError:
    print(f'Directory {prefix_path} already exists, continuing script.')

workbook_path = f'template_DasFacts.xlsx' # Path of workbook template. 
work_book = openpyxl.load_workbook(workbook_path)
sheet = work_book['Sheet1']
cursor = connection.cursor()

def GetSystemParameters():
    cursor.execute('SELECT SYS_PARAMS.PARAMETER_INDEX, SYS_PARAMS.PARAMETER_NAME, SYS_PARAMS.SITE, SYS_PARAMS.FORMULA, SYS_PARAMS.LOWER_LIMIT_CU, SYS_PARAMS.UPPER_LIMIT_CU, SYS_PARAMS.LOWER_LIMIT_EU, SYS_PARAMS.UPPER_LIMIT_EU, SYS_PARAMS.UNIT_CONVERSION, SYS_PARAMS.DATA_SOURCE, ALM_CONFIG.LOW_LIMIT_IND, ALM_CONFIG.HI_LIMIT_IND ,ALM_CONFIG.HIHI_LIMIT_IND FROM SYS_PARAMS LEFT JOIN ALM_CONFIG on SYS_PARAMS.PARAMETER_INDEX=ALM_CONFIG.PARAMETER_INDEX')
    sysParamData = []
    for row in cursor.fetchall():  

        splitString = row[1].split(',')
        if(len(splitString) > 1): 
            splitString = splitString[1]
        else: 
            splitString = 'None'

        data_source = ''

        if(row[9] == 0):
            data_source = 'Spare'

        if(row[9] == 1):
            data_source = 'Historical'

        if(row[9] == 2):
            data_source = 'Monitored'

        if(row[9] == 3): 
            data_source = 'Calculated' 

        sqlRow = { 
            # Adding the correct SQL Database field into our dictionary. 
            # FORMAT:
            # Key in the dictionary: SQL Database Column data.
            "param_index": row[0],
            "param_name": row[1],
            "param_units": splitString,
            "site": row[2],
            "calc_id": row[3],
            "lower_limit_cu": row[4],
            "upper_limit_cu": row[5],
            "lower_limit_eu": row[6],
            "upper_limit_eu": row[7],
            "unit_conversion": row[8],
            "data_source": data_source,
            "low_limit": row[10],
            "hi_limit": row[11],
            "hi_hi_limit": row[12]
        }
        sysParamData.append(sqlRow)
    return sysParamData

# Create Sys_Param List
params_to_export = GetSystemParameters()

# Iterate through Sys_Params list.
for parameter in range(len(params_to_export)):
    for db_field in range(len(params_to_export[parameter])):
        column_letter = get_column_letter(db_field+1)
        sheet[f'{column_letter}{parameter+2}'] = list(params_to_export[parameter].values())[db_field] # We use x+2 as the index to start AFTER Row 1. 

try:
    work_book.save(f'{prefix_path}{database_name}DasFacts.xlsx')
    print(f'{prefix_path}{database_name}DasFacts.xlsx has been created')
except Exception as err: # Writes error message to log file and exits the script.
    error_string = f'''ERROR: An error occurred while saving the {database_name}DasFacts.xlsx. It is likely that it already existed and was open in Excel. Close Excel and try again. \n\n An error log has been created in: \n{prefix_path}\n'''
    print(error_string)
    log = open(f'{prefix_path}das_facts.log', 'a')
    log.write(f'{str(datetime.datetime.now())}\n')
    log.write('--------------------------------------\n')
    log.write(f'{error_string}\n\n')
    
connection.close()