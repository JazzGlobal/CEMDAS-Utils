# Queries the SYS_PARAMS table of a specified database
# Exports the results to a spreadsheet to generate the DasFactsSheet

import sys, os, openpyxl, pymssql
from openpyxl.utils import get_column_letter

database = input('Enter the name of the Sys1 database (Without Sys1)')
conn = pymssql.connect(
                server=f"{os.getlogin()}\\SQLEXPRESS",
                port=1433,
                user="sa",
                password="sa123",
                database=(f'{database}Sys1'))

prefix_path = f'C:\\Users\\{os.getlogin()}\\Desktop\\Generated_Scripts\\'
worksheet_path = f'{prefix_path}template_DasFacts.xlsx'
wb = openpyxl.load_workbook(worksheet_path)
sheet = wb['Sheet1']
cursor = conn.cursor()
cursor.execute('SELECT * FROM SYS_PARAMS')

def queryDatabase():
    sqlData = []
    for row in cursor.fetchall():  
        dataRow = {
            "param_index": row[0],
            "param_name": row[1],
            "site": row[2],
            "calc_id": row[13],
            "lower_limit_cu": row[23],
            "upper_limit_cu": row[24],
            "lower_limit_eu": row[25],
            "upper_limit_eu": row[26],
            "unit_conversion": row[39],
            "data_source": row[40],
        }
        sqlData.append(dataRow)
    return sqlData

toExport = queryDatabase()

for x in range(len(toExport)):
    for y in range(len(toExport[x])):
        column_letter = get_column_letter(y+1)
        sheet[f'{column_letter}{x+2}'] = list(toExport[x].values())[y]

wb.save(f'{prefix_path}newDasFacts.xlsx')
conn.close()